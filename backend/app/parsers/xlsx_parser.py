# -*- coding: utf-8 -*-
"""
app/parsers/xlsx_parser.py —— Excel 解析器（openpyxl）

- 每个工作表 → 标题块（sheet 名，level 1）+ 表格块（整表行数据）
- page 语义 = sheet 序号（xlsx 无"页"概念；出处锚点以此为准，如实记录）
- read_only 模式低内存；行数上限 2000 防爆；空行跳过
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..schemas import Block, BlockType, ParsedDocument
from .base import build_section_tree

MAX_ROWS_PER_SHEET = 2000


class XlsxParser:
    file_type = "xlsx"

    def parse(self, path: Path | str) -> ParsedDocument:
        path = Path(path)
        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            blocks: list[Block] = []
            char_count = 0
            bid = [0]

            def new_block(type_, text, page, table=None, level=None) -> Block:
                bid[0] += 1
                return Block(block_id=f"B{bid[0]:04d}", type=type_, text=text,
                             page=page, table=table, level=level)

            for sheet_no, ws in enumerate(wb.worksheets, start=1):
                blocks.append(new_block(
                    BlockType.HEADING, f"工作表：{ws.title}", sheet_no, level=1))
                rows: list[list[str]] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= MAX_ROWS_PER_SHEET:
                        break
                    cells = ["" if c is None else str(c).strip() for c in row]
                    if not any(cells):
                        continue
                    rows.append(cells)
                if rows:
                    text = "\n".join(" | ".join(r) for r in rows)
                    blocks.append(new_block(BlockType.TABLE, text, sheet_no, table=rows))
                    char_count += len(text)

            sections = build_section_tree(blocks)
            return ParsedDocument(
                schema_version="1.0.0",
                file_name=path.name,
                file_type="xlsx",
                total_pages=len(wb.worksheets),   # page 语义 = sheet 序号
                char_count=char_count,
                sections=sections,
                blocks=blocks,
            )
        finally:
            wb.close()
